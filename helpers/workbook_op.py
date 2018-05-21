from openpyxl import load_workbook
import numpy as np
import os


def open_workbook(filename):
    wb = load_workbook(filename=filename)
    st_names = wb.sheetnames


    st_name_dict = {}
    for i in range(len(st_names)):
        st_name_dict[i] = st_names[i]
    return wb, st_name_dict

def parse_sheets(wb, selected, st_name_dict):
    all_content = None
    for i in selected:
        name = st_name_dict[i]
        st = wb[name]
        content = parse_sheet(st, i)
        if all_content is None:
            all_content = content
        else:
            all_content = np.vstack((all_content, content))
    perm = np.random.permutation(all_content.shape[0])
    return all_content[perm, :]


def parse_sheet(st, st_number):

    max_row = st.max_row

    from_st = []
    words = []
    meanings = []
    synonyms = []
    reports = []
    accuracies = []

    def append_row_to_lists(row_idx):

        wd = str(st.cell(row=row_idx, column=1).value)
        mg = str(st.cell(row=row_idx, column=2).value)
        sy = str(st.cell(row=row_idx, column=3).value)
        re = st.cell(row=row_idx, column=4).value
        ac = st.cell(row=row_idx, column=5).value
        if wd == 'None':
            return False

        if len(wd) == 0 or len(mg) == 0:
            raise ValueError('Empty Found in line {}'.format(row_idx))


        from_st.append(st_number)
        words.append(wd.strip())
        meanings.append(mg.strip())
        synonyms.append(sy.strip())
        reports.append('' if re is None else re.strip())
        accuracies.append('' if ac is None else ac.strip())

        return True

    for i in range(max_row):
        row_idx = i + 1
        re = append_row_to_lists(row_idx)
        if re == False: break

    content = np.array([from_st, words, meanings, synonyms, reports, accuracies], dtype=object).T
    '''
    [sheet ID][word][meaning][synonyms][reports][accuracies]
    [6]       [good][好]      [well]   [10/10]   [100%]
    '''
    return content


def fill_sheet_with_new_content(st, content):
    ct_max_row = content.shape[0]
    st_max_row = st.max_row
    assert ct_max_row <= st_max_row, "content row: {} > sheet row: {} | at sheet {}".format(ct_max_row, st_max_row, st.title)
    for r in range(ct_max_row):
        st.cell(row=r + 1, column=4, value=content[r , -2])
        st.cell(row=r + 1, column=5, value=content[r , -1])


def save_workbook(wb, path):
    overwrite = True
    if os.path.isfile(path):
        inp = input("{} already exist, would you like to overwrite it? [y/N]".format(path)).strip().lower()
        if inp == 'y':
            overwrite = True
        else:
            overwrite = False

    if overwrite:
        wb.save(filename=path)
        return path
    else:
        while True:
            inp  = input("Please input a path to save or input -1 to exit:").strip()
            if len(inp) <=0:
                continue

            try: inpp = int(inp)
            except: inpp = None
            if inpp is not None:
                if inpp == -1:
                    print("Exit without saving, Bye :D")
                    exit()
                else:
                    continue
            else:
                if os.path.isdir(os.path.dirname(inp)):
                    if inp.endswith('.xlsx'):
                        wb.save(filename=inp)
                        return inp
                    else:
                        print('Input path should end with .xlsx')
                        continue
                else:
                    print("no such a dir: {}".format(os.path.dirname(inp)))
                    continue


